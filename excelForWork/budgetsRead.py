import openpyxl

def getKakiat():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/Kakiat Budget 2020 ERCSD- principal budget.xlsx', data_only=True)

    sheet = wb['Sheet1']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Aggregate'
    row2 = 1
    #Technology Hardware
    for i in range(31,38):
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            


    #Technology Software
    for i in range(70,81):
        print("\nBegin software\n")
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    
    wb2.save('C:/Users/aj282/kakiat.xlsx')

def getSvhs():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/SVHS-ERCSD-budget-principal_form.xlsx', data_only=True)

    sheet = wb['Sheet1']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Aggregate'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(57,137):
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            


    #Technology Software
    for i in range(140,144):
        print("\nBegin software\n")
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    
    wb2.save('C:/Users/aj282/svhs.xlsx')

# Do not copy this formatting. RHS had significant formatting errors and is not replicable to other data sheets.
def getRhs():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/Ramapo Budget Principal 2020-2021.xlsx', data_only=True)

    sheet = wb['DO NOT MAKE ADDITIONAL SHEETS']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Aggregate'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(44,70):
        rowNum = i
        cell = 'B' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'H' + str(rowNum)
            cost = 'I' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            


    #Technology Software
    for i in range(72,86):
        print("\nBegin software\n")
        rowNum = i
        cell = 'F' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'H' + str(rowNum)
            cost = 'I' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    
    wb2.save('C:/Users/aj282/rhs.xlsx')

def getEld():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/Eldorado Budget.xlsx', data_only=True)

    sheet = wb['Sheet1']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Aggregate'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(17,36):
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            

    '''
    #Technology Software
    for i in range(140,144):
        print("\nBegin software\n")
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    '''  
    wb2.save('C:/Users/aj282/eld.xlsx')

def getElm():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/ELMWOOD-Budget Principal Form 2020-2021 FINAL.xlsx', data_only=True)

    sheet = wb['PRINCIPAL']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Aggregate'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(19,23):
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            


    #Technology Software
    for i in range(28,30):
        print("\nBegin software\n")
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    
    wb2.save('C:/Users/aj282/elm.xlsx')

def getPom():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/Pomona Principal Budget Form_1.xlsx', data_only=True)

    sheet = wb['Teacher']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Aggregate'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(18,26):
        rowNum = i
        cell = 'A' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            

    '''
    #No software listed
    #Technology Software
    for i in range():
        print("\nBegin software\n")
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    '''
    wb2.save('C:/Users/aj282/pom.xlsx')

def getSpk():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/Summit Park-ERCSD-Budget Principal Form 2020-2021 DW.xlsx', data_only=True)

    sheet = wb['Sheet1']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Sheet1'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(81,122):
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            


    #Technology Software
    for i in range(127,138):
        print("\nBegin software\n")
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    
    wb2.save('C:/Users/aj282/spk.xlsx')

def getMargetts():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/Margetts Principals Resource Request Form 2020.xlsx', data_only=True)

    sheet = wb['Sheet1']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Sheet1'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(19,25):
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            

    '''
    #Formatted to poorly to munge useful data
    #Technology Software
    for i in range(127,138):
        print("\nBegin software\n")
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    '''
    
    wb2.save('C:/Users/aj282/mar.xlsx')

def getLkn():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/Lime Kiln-ERCSD-Budget Principal Form 2020-2021.xlsx', data_only=True)

    sheet = wb['Sheet1']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Sheet1'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(28,54):
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            
    '''
    #Technology Software
    for i in range(127,138):
        print("\nBegin software\n")
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    '''

def getHemp():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/Hempstead budget-principal form 2020.xlsx', data_only=True)

    sheet = wb['Sheet1']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Sheet1'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(17,25):
        rowNum = i
        cell = 'A' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            
    
    #Technology Software
    for i in range(26,30):
        print("\nBegin software\n")
        rowNum = i
        cell = 'A' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    
    wb2.save('C:/Users/aj282/hemp.xlsx')

def getHemp():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/Hempstead budget-principal form 2020.xlsx', data_only=True)

    sheet = wb['Sheet1']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Sheet1'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(17,25):
        rowNum = i
        cell = 'A' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            
    
    #Technology Software
    for i in range(26,30):
        print("\nBegin software\n")
        rowNum = i
        cell = 'A' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    
    wb2.save('C:/Users/aj282/hemp.xlsx')

def getFle():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/Fleetwood-Budget-principal_form 2019-2020.xlsx', data_only=True)

    sheet = wb['Sheet1']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Sheet1'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(17,27):
        rowNum = i
        cell = 'E' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            
    
    #Technology Software
    for i in range(29,34):
        print("\nBegin software\n")
        rowNum = i
        cell = 'A' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    
    wb2.save('C:/Users/aj282/fle.xlsx')

def getGrandview():
    wb = openpyxl.load_workbook('C:/Users/aj282/Downloads/School budget proposals/School budget proposals/Grandview-budget-principal_form 2020-2021.xlsx', data_only=True)

    sheet = wb['Sheet1']

    #sheetsDict = {}

    rowNum = 1

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = 'Sheet1'
    row2 = 1

    #Technology Hardware (Note: Other hardware/bad data to be filtered)
    for i in range(15,20):
        rowNum = i
        cell = 'A' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
            
    '''
    #No software listed  
    #Technology Software
    for i in range(29,34):
        print("\nBegin software\n")
        rowNum = i
        cell = 'A' + str(rowNum)
        if sheet[cell].value == None:
            #print("Empty cell: {}".format(cell))
            continue
        else:
            print()
            quant = 'G' + str(rowNum)
            cost = 'H' + str(rowNum)
            value = str(sheet[quant].value) + " ($" + str(sheet[cost].value) + ")"
            print("{}: {}".format(cell, sheet[cell].value))
            row2 += 1
            sheet2CellA = 'A' + str(row2)
            sheet2CellB = 'B' + str(row2)
              
            if sheet[cell].value.find(":") != -1:
                print("{}: {}".format(cell, sheet[cell].value[:sheet[cell].value.find(":")]))
                sheet2[sheet2CellA] = sheet[cell].value[:sheet[cell].value.find(":")]
                sheet2[sheet2CellB] = value
            else:
                print("{}: {}...{}".format(cell, sheet[cell].value[:30], value))
                sheet2[sheet2CellA] = sheet[cell].value[:30]
                sheet2[sheet2CellB] = value
    '''
    wb2.save('C:/Users/aj282/grand.xlsx')




if __name__ == '__main__':
    #getKakiat()
    #getSvhs()
    #getRhs()
    #getEld()
    #getElm()
    #getPom()   
    #getSpk()
    #getMargetts()
    #getLkn()
    #getHemp()
    #getFle()
    getGrandview()







